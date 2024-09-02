import openpyxl
from openpyxl import Workbook

def split_excel_file(input_file, output_file_prefix, num_splits):
    # Load the input workbook
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    
    # Get total number of rows
    total_rows = ws.max_row
    
    # Calculate number of rows per split
    rows_per_split = total_rows // num_splits
    
    # If there are remaining rows, we need to distribute them
    remaining_rows = total_rows % num_splits
    
    current_row = 1
    
    for i in range(num_splits):
        # Create a new workbook for each split
        new_wb = Workbook()
        new_ws = new_wb.active
        
        # Copy the header from the original file
        for col in ws.iter_cols(1, ws.max_column):
            new_ws.cell(row=1, column=col[0].column, value=col[0].value)
        
        # Calculate the number of rows for this split
        split_rows = rows_per_split + (1 if i < remaining_rows else 0)
        
        # Copy rows from the original worksheet to the new worksheet
        for row_index in range(1, split_rows + 1):
            for col_index, cell in enumerate(ws.iter_cols(1, ws.max_column), start=1):
                new_ws.cell(row=row_index + 1, column=col_index, value=ws.cell(row=current_row + 1, column=col_index).value)
            current_row += 1
        
        # Save the new split workbook
        output_file = f"{output_file_prefix}_{i+1}.xlsx"
        new_wb.save(output_file)
        print(f"Saved {output_file} with {split_rows} rows.")

# Usage
input_file = 'Splitting_Data.xlsx'  # Replace with your input file path
output_file_prefix = 'split_file'    # Replace with your desired output file prefix
num_splits = 14                      # Number of splits

split_excel_file(input_file, output_file_prefix, num_splits)
